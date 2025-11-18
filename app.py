import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from PIL import Image
import cv2
import numpy as np
from pyzxing import BarCodeReader

# Colores
COLOR_VERDE = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
COLOR_MORADO = PatternFill(start_color="800080", end_color="800080", fill_type="solid")

st.title("📚 Inventario Biblioteca")
st.write("Escanea códigos con la cámara del celular. Si el código existe, se marcará en verde; si es nuevo, se agregará en morado.")

# Subir archivo Excel
uploaded_file = st.file_uploader("Sube tu archivo Excel del inventario", type=["xlsx"])
if uploaded_file:
    excel_path = "inventario.xlsx"
    with open(excel_path, "wb") as f:
        f.write(uploaded_file.getbuffer())

    wb = load_workbook(excel_path)
    sheet = wb.active
    df = pd.read_excel(excel_path)

    # Detectar columna de códigos
    codigo_columna = None
    for col in df.columns:
        if "codigo" in col.lower():
            codigo_columna = col
            break

    if not codigo_columna:
        st.error("No se encontró ninguna columna que contenga 'codigo'.")
    else:
        codigo_a_fila = {str(row[codigo_columna]): idx+2 for idx, row in df.iterrows()}

        st.subheader("Escanea el código")
        img_file = st.camera_input("Toma una foto del código")

        if img_file:
            img = Image.open(img_file)
            img_np = np.array(img)

            # Convertir imagen para OpenCV
            img_cv = cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)

            # Usar pyzxing para decodificar
            reader = BarCodeReader()
            temp_path = "temp_image.png"
            img.save(temp_path)
            result = reader.decode(temp_path)

            if result and len(result) > 0:
                codigo = result[0]['parsed']
                st.write(f"✅ Código detectado: **{codigo}**")

                if codigo in codigo_a_fila:
                    fila = codigo_a_fila[codigo]
                    celda = f"A{fila}"
                    sheet[celda].fill = COLOR_VERDE
                    sheet[celda].font = Font(bold=True)
                    st.success(f"Código {codigo} encontrado y marcado en verde.")
                else:
                    nueva_fila = sheet.max_row + 1
                    sheet[f"A{nueva_fila}"] = codigo
                    sheet[f"A{nueva_fila}"].fill = COLOR_MORADO
                    sheet[f"A{nueva_fila}"].font = Font(bold=True)
                    st.warning(f"Código {codigo} agregado como nuevo y marcado en morado.")

                wb.save(excel_path)
            else:
                st.error("No se detectó ningún código en la imagen.")

        st.subheader("Inventario actualizado")
        st.dataframe(pd.read_excel(excel_path))

        with open(excel_path, "rb") as f:
            st.download_button("Descargar Excel actualizado", f, file_name="inventario_actualizado.xlsx")
